# tools/inspect_excel.py
import os
import sys
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(__file__))
path = os.path.join(ROOT, 'program_export_sample.xlsx')
if not os.path.exists(path):
    print('File not found:', path)
    sys.exit(1)

wb = load_workbook(path, data_only=True)
print('Workbook:', path)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    print('\n--- Sheet:', sheet)
    for r in ws.iter_rows(values_only=True):
        # print non-empty rows compactly
        if not any(c is not None and c != '' for c in r):
            continue
        print(r)

print('\nDone')
