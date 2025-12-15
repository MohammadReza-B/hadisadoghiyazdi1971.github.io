# tools/generate_excel_from_plan.py
# Generates an Excel workbook from the CURRENT_PLAN in app.config using the same logic as exportToExcel
import os, sys, json
ROOT = os.path.dirname(os.path.dirname(__file__))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from app import create_app
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

app = create_app()

with app.app_context():
    # If CURRENT_PLAN is missing, seed the DATASET and CURRENT_PLAN with the example from the user
    jobs = [
        {'id': 'job-017', 'address': 'مشهد، بلوار هاشمیه، پلاک ۱۲۳ - شماره 17', 'lat': 36.319542, 'lon': 59.605757},
        {'id': 'job-019', 'address': 'مشهد، بلوار سجاد، مجتمع تجاری - شماره 19', 'lat': 36.326162, 'lon': 59.60086},
        {'id': 'job-014', 'address': 'مشهد، قاسم آباد، خیابان شهیدان - شماره 14', 'lat': 36.303664, 'lon': 59.608943},
        {'id': 'job-018', 'address': 'مشهد، کوهسنگی، خیابان بهشت - شماره 18', 'lat': 36.339166, 'lon': 59.60396}
    ]
    groups = [
        {'id': 'teamA', 'label': 'تیم الف', 'members_count': 3, 'daily_capacity_hours': 18, 'skills': [], 'base_lat': 36.317054, 'base_lon': 59.605757}
    ]

    app.config['DATASET']['jobs'] = jobs
    app.config['DATASET']['groups'] = groups

    plan = app.config.get('CURRENT_PLAN') or {
        'plan': {
            'teamA': {
                'Saturday': [
                    {'job_id': 'job-017', 'start': 8.05, 'duration': 1.1},
                    {'job_id': 'job-019', 'start': 9.1667, 'duration': 2},
                    {'job_id': 'job-014', 'start': 11.2333, 'duration': 2.1}
                ],
                'Sunday': [
                    {'job_id': 'job-018', 'start': 8.0, 'duration': 2.7}
                ]
            }
        }
    }
    # ensure the app's CURRENT_PLAN is set so /api/map-data/batch can read it
    app.config['CURRENT_PLAN'] = plan

    wb = Workbook()
    ws = wb.active
    ws.title = 'params'
    # minimal params sheet
    ws.append(['پارامترها'])

    def find_job(jid):
        return next((j for j in jobs if j['id'] == jid), {})

    # prepare batch request: collect all group/day pairs
    batch_items = []
    for gid in plan['plan'].keys():
        for day in plan['plan'][gid].keys():
            if plan['plan'][gid][day]:
                batch_items.append({'group': gid, 'day': day})

    # call batch endpoint once
    import json as _json
    batch_map = {}
    if batch_items:
        payload = _json.dumps({'items': batch_items})
        headers = {'Content-Type': 'application/json'}
        with app.test_request_context('/api/map-data/batch', method='POST', data=payload, headers=headers):
            resp = app.full_dispatch_request()
            resp_json = _json.loads(resp.get_data(as_text=True))
        for it in resp_json.get('items', []):
            key = f"{it.get('group')}|||{it.get('day')}"
            batch_map[key] = it

    for gid in plan['plan'].keys():
        group = next((c for c in groups if c['id'] == gid), {'label': gid})
        ws = wb.create_sheet(title=group.get('label') or gid)
        # replicate basic structure: for each day, write rows
        for day in plan['plan'][gid].keys():
            assignments = plan['plan'][gid][day]
            if not assignments:
                continue
            ws.append([f'روز: {day}'])
            ws.append(['کار','آدرس','اولویت','ساعت شروع','مدت (ساعت)','نفر-ساعت','موقعیت جغرافیایی','مسافت طی شده (کیلومتر)'])

            batch_key = f"{gid}|||{day}"
            item = batch_map.get(batch_key, {})
            route_jobs = item.get('jobs', []) or []
            route_total = item.get('route_total_km')

            # virtual row
            if route_jobs:
                first = route_jobs[0]
                coords = f"{first.get('lat')}, {first.get('lon')}" if first.get('lat') and first.get('lon') else ''
                ws.append(['پایگاه → اولین مقصد', first.get('address',''), '', '', '', '', coords, first.get('route_distance_km')])

            # job rows
            for idx, a in enumerate(assignments):
                job = find_job(a['job_id'])
                coords = f"{job.get('lat')}, {job.get('lon')}" if job.get('lat') and job.get('lon') else ''
                # distance logic same as JS: leave first job empty, subsequent jobs take route_jobs[idx].route_distance_km
                dist = ''
                if route_jobs:
                    if len(route_jobs) == 1:
                        dist = ''
                    else:
                        if idx > 0 and len(route_jobs) > idx:
                            dist = route_jobs[idx].get('route_distance_km')
                ws.append([a['job_id'], job.get('address',''), job.get('priority',''), a.get('start'), a.get('duration'), '', coords, dist])

            # summary rows
            # compute sum of last column for this day's block
            # find how many rows we just added: virtual row (1 if present) + len(assignments)
            block_rows = len(assignments) + (1 if route_jobs else 0)
            # sum from the last block_rows rows
            sum_km = 0.0
            start_row = ws.max_row - block_rows + 1
            for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=8, max_col=8):
                for cell in row:
                    val = cell.value
                    if isinstance(val, (int, float)):
                        sum_km += float(val)
            ws.append([])
            ws.append(['', '', '', '', '', '', 'جمع مسافت بخش‌ها (km):', sum_km])
            ws.append(['', '', '', '', '', '', 'مجموع مسیر OSRM (km):', route_total])
            ws.append([])

    out_path = os.path.join(ROOT, 'program_export_sample.xlsx')
    wb.save(out_path)
    print('Wrote', out_path)
